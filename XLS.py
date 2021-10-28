import calendar
import datetime
from openpyxl.workbook import Workbook
from openpyxl.styles import Font, Alignment
from openpyxl.utils import get_column_letter
import data
import database
from openpyxl import load_workbook


def month_to_csv(year=2021, month=1):
    """
    :param year: pobiera rok
    :param month: pobiera wartość miesiąca od 1 do 12
    :returns: zwraca plik z tabelą z podanym miesiącem
    """
    wb = Workbook()
    ws = wb.active
    list_of_days = []
    cal1 = calendar.Calendar()

    for day in cal1.itermonthdays2(year, month):
        if day[0] == 0:
            continue
        list_of_days.append(day)

    ws.row_dimensions[1].height = 50

    i = 0
    while i != len(list_of_days):
        ws.cell(column=i + 3, row=2,
                value=str(datetime.date(year, month, list_of_days[i][0]).day)).alignment = Alignment(
            horizontal='center', vertical='center')
        ws.cell(column=i + 3, row=2).font = Font(size=10)
        if list_of_days[i][1] == 5 or list_of_days[i][1] == 6:
            ws.cell(column=i + 3, row=2).font = Font(color="FF0000", bold=True)
            ws.cell(column=i + 3, row=2).alignment = Alignment(horizontal='center', vertical='center')

        ws.column_dimensions[get_column_letter(i + 3)].width = 5
        i += 1
    ws.merge_cells(start_row=1, start_column=3, end_row=1, end_column=len(list_of_days) + 2)
    ws.cell(column=3, row=1, value='{} {}'.format(year, data.month_list[month - 1])).font = Font(bold=True, size=20)
    ws.cell(column=3, row=1).alignment = Alignment(horizontal='center', vertical='center')

    database.my_cursor.execute("SELECT * FROM Persons")
    workers = database.my_cursor.fetchall()
    for i in range(len(workers)):
        # j = i
        a = int(3 + 7 * i)
        b = int(8 + 7 * i)
        merge_a = 'A{}:A{}'.format(a, b)
        merge_b = 'B{}:B{}'.format(a, b)
        ws.merge_cells(merge_a)
        ws.merge_cells(merge_b)
        ws.cell(column=1, row=a, value=workers[i][0]).alignment = Alignment(
            horizontal='center', vertical='center')
        ws.column_dimensions[get_column_letter(1)].width = 5
        ws.column_dimensions[get_column_letter(1)].bestFit = True
        ws.cell(a, 2, value='{} {}'.format(workers[i][1], workers[i][2]))
        ws.cell(a, 2).alignment = Alignment(horizontal='center', vertical='center')
        ws.column_dimensions[get_column_letter(2)].width = 20
    return wb

# def workers_in_worksheet(file_name):
#     wb = load_workbook(filename=file_name)
#     ws = wb.active
#     database.my_cursor.execute("SELECT * FROM Persons")
#     workers = database.my_cursor.fetchall()
#     for i in range(len(workers)):
#         ws.merge_cells(start_row=3 + 6 * i, start_column=1, end_row=9 + 6 * i, end_column=1)
#         ws.cell(column=1, row=3 + 6 * i, value="{} {}".format(workers[i][1], workers[i][2])).alignment = Alignment(
#             horizontal='center', vertical='center')
#     return wb
