"""
Projekt ma za zadanie pobranie plików wsadowych (czyt. godziny oraz prośby) zaczytanie ich, a następnie
zwrócenie danych w postaci tabeli miesięcznej na zadany rok oraz miesiąc z informacją kto i kiedy ma dyżur, łącznie z
uwzględnieniem dni wolnych, wolnych dni po 24h dyżuru oraz tak żeby obsadzone były wszystkie dni i godziny w
wystarczającej ilości osób.
"""

import XLS
from employers import Employers
from openpyxl import load_workbook
from openpyxl import formatting
from data import month_list, path
import os
from openpyxl import writer


def add_work_schedule():
    year = int(input("Podaj rok: "))
    month = int(input("Podaj miesiąc: "))
    if not os.path.isfile(path + '{} {}.xlsx'.format(year, month_list[month - 1])):
        XLS.month_to_csv().save('{} {}.xlsx'.format(year, month_list[month - 1]))
    else:
        print("taki grafik już istnieje!")
    pass

def add_worker():
    x = input("Podaj imię pracownika: ")
    y = input("Podaj nazwisko pracownika: ")
    worker = Employers(x, y)
    worker.add_worker()
    pass


add_worker()
add_work_schedule()







# def open_worbook(path):
#     workbook = load_workbook(filename=path)
#     print("Worksheet names: {}".format(workbook.sheetnames))
#     sheet = workbook.active
#     print(sheet)
#     print("The title of the Worksheet is: {}".format(sheet.title))


# def employer():
#     return


# xls.save(file_name)

# open_worbook(file_name)
